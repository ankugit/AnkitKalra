
# coding: utf-8

# In[59]:


#Reading the Excel sheet
from itertools import combinations, permutations
import pandas as pd
import random
import numpy as np
time_remove=float(input("Time taken to  remove the component from the channel(in min): "))
time_load=float(input("Time taken to load the component on the channel(in min): "))

channel=[0,0]
excel_file="data.xlsx"
data=pd.read_excel(excel_file)
xl=pd.ExcelFile(io="data.xlsx")
data.head()


#Number of PCBs
PCB=data.shape[1]
Total_Component=data.shape[0]
print("Total kinds of PCBs: {}".format(PCB))
print("Total kinds of components: {}".format(Total_Component))

def PCB_list(data):
    temp=[]
    for i in data.head():
        temp.append(i)
    return temp
pcb_list=PCB_list(data)
pcb_list

#The name of the PCB is passed and requirements are returned in a list format
def comp(t):
    temp=[]
    for j in data.index:
        if data[t][j]==1:
            temp.append(j)
    return temp
#number of ways to make PCBs
def ways_to_make_pcb(pcb_list):
    temp=list(permutations(pcb_list))
    print(len(temp))
    return temp

def init_channel(channel):
    #from the first combination it chooses any PCB at random
    start_pcb=(comb_pcb[0][0])
    #the components of the random pcb are taken
    start_pcb_comp=d[start_pcb]
    #The channel is set according to the requirements of the first PCB
    #The problem might occur if the pcb requires only one PCB
    
    #If the PCb requires only one component the first channel is set according to the first pcb and the second one
    #is selected at random from the next pcb to be made
    global time
    time=0
    if len(d[start_pcb])==1:
        channel[0]=random.choice(start_pcb_comp)
        time=time+time_load
        #print(time)
        while True:
            channel[1]=random.choice(d[comb_pcb[0][1]])
            if channel[0]!=channel[1]:
                break
        time=time+time_load
        #print(time)
    #In case the first pcb requires atleast 2 PCB's, we will not face any problem
    else:
        channel[0]=random.choice(start_pcb_comp)
        time=time+time_load
        #print(time)
        while True:
            channel[1]=random.choice(start_pcb_comp)
            if(channel[0]!=channel[1]):
                time=time+time_load
                #print(time)
                break
        
    #print("Channel-1:{}\nChannel-2:{}".format(channel[0], channel[1]))
    
#The current pcb and the current combination is passed into the function and returns the next pcb to be made
def next_pcb(current_pcb, current_comb):
    return current_comb[current_comb.index(current_pcb)+1]


#After placing the components pass this function to check if the pcb is made or not...if not
#the list of remaining components will be passed

#Take note here the temp stores the components placed on the PCB....inorder to make the current_pcb i had to change the para
#meter to current_pcb_temp as the names were clashing
def check(current_pcb_temp, channel, current_comb, temp_temp):
    #store the components placed after checking if they are required or not
    #try:
    global temp
    temp=temp_temp
    global current_pcb
    current_pcb=current_pcb_temp
    if channel[0] in d[current_pcb]:
        temp.append(channel[0])
    if channel[1] in d[current_pcb]:
        temp.append(channel[1])
    #inorder to avoid duplications in the list
    temp=list(set(temp))  
    #except:
        #try:
            #if d[current_pcb].index(channel[1]):
                #temp.append(channel[1])
        #except:
            #print("")
    #Temp stores the items placed on the PCB, if the pcb is made the "PCB made" is flashed and the current_pcb is 
    #changed to the next PCB in the combination
    if len(np.setdiff1d(d[current_pcb], temp, assume_unique=False)) ==0:
        #print("PCB made- {}".format(current_pcb))
        #print(current_pcb)
        current_pcb=next_pcb(current_pcb, current_comb)
        #print(current_pcb)
        #Once the PCB is made no need for the temp
        #print("The requirement is met", temp)
        temp=[]
        modify_channel(current_pcb, d[current_pcb], channel, current_comb, temp)
    else:
        modify_channel(current_pcb,np.setdiff1d(d[current_pcb], temp), channel, current_comb,temp)

#The current pcb, the required components to complete the current pcb , the channel config and the current comb is passed 
#If only 1 comp is left to make the pcb, the channel is checked if it has that or not
    #If it has then the other channel is updated according to the next PCB to be made
#If more than 1 components are required then both the channels are updated as per the requirements
def modify_channel(current_pcb,comp_rem, channel, current_comb,temp):
    global time
    if len(comp_rem)==1:
        #try block required if channel.index() gives an error
        try:
            if channel.index(comp_rem)==0:
                channel[1]=random.choice(d[next_pcb(current_pcb, current_comb)])
                time=time+time_remove+time_load
                #print(time)
            else:
                channel[0]=random.choice(d[next_pcb(current_pcb, current_comb)])
                time=time+time_remove+time_load
                #print(time)
        except:
            channel[0]=comp_rem[0]
            #print(time)
            while True:
                channel[1]=random.choice(d[next_pcb(current_pcb, current_comb)])
                if channel[0]!=channel[1]:
                    time=time+time_remove+time_load
                    #print(time)
                    break
       # current_pcb=next_pcb(current_pcb, current_comb)
    else:
        channel[0]=random.choice(comp_rem)
        #print(time)
        while True:
            channel[1]=random.choice(comp_rem)
            if channel[0]!=channel[1]:
                time=time+time_remove+time_load
                #print(time)
                break
        
        

#Making a dictionary where we will store the requirements of PCBs
d={}
for i in range(len(pcb_list)):
    d[pcb_list[i]]=comp(pcb_list[i])
#Number of ways to make PCBs
comb_pcb=ways_to_make_pcb(pcb_list)

#Taking input from the user to give the input times


import xlsxwriter
workbook=xlsxwriter.Workbook("test_1.xlsx")
worksheet=workbook.add_worksheet()


for j in range(100):
    for i in range(len(comb_pcb)):
        init_channel(channel)
        current_pcb=0
        current_pcb=comb_pcb[i][0]
        current_comb=comb_pcb[i]
        #the channel has been init.
        #print(channel, current_pcb, current_comb)
        #print("Initial PCB in starting", current_pcb)

        if j==0:
            temp=[]
            while True:
                try:
                    #print("Current channel before starting",channel)
                    check(current_pcb, channel, current_comb, temp)
                    #print("channel after first mod.",channel)
                    #print("Current PCB: ",current_pcb)
                except:
                    break
            worksheet.write(i, 0, str(current_comb))
            worksheet.write(i, j+1, time)
            #print("Current Comb: {}------ Time: {}".format(current_comb, time))
        else:
            temp=[]
            while True:
                try:
                    #print("Current channel before starting",channel)
                    check(current_pcb, channel, current_comb, temp)
                    #print("channel after first mod.",channel)
                    #print("Current PCB: ",current_pcb)
                except:
                    break
            worksheet.write(i, 0, str(current_comb))
            worksheet.write(i, j+1, time)
            #print("Current Comb: {}------ Time: {}".format(current_comb, time))
            
        
        
workbook.close()
    


# In[6]:


file=open("data.txt", "w")
file.write("Hello")
file.close()
file=open("data.txt", "a")
file.write("HELLLLOO")
file.close()


# In[7]:


from openpyxl import Workbook
wb=Workbook()


# In[8]:


ws=wb.active


# In[9]:


ws.title="Test-1"


# In[10]:


ws.cell(row=1, column=1, value=10)
wb.save("test.xlsx")

